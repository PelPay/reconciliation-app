"""
Streamlit app — Pelpay ↔ Gateway Settlement Reconciliation

Upload all files — auto-detects Pelpay, Cybersource, ChoicePay, and MPGS.

Usage:
    python -m streamlit run app.py
"""
import sys, os, tempfile, csv, io
from datetime import date
from collections import defaultdict
import streamlit as st
import openpyxl

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from reconcile_core import run, DEFAULT_SCHEMAS

st.set_page_config(page_title='Reconciliation App', layout='wide')
st.title('Gateway ↔ Pelpay Settlement Reconciliation')

# ── Detection ────────────────────────────────────────────
def classify_file(headers):
    """Return ('PELPAY', None, None), ('SETTLEMENT', gateway, currency), or None."""
    h = [str(c).strip().lower() for c in headers]
    has = lambda s: any(s in col for col in h)

    # Pelpay: has Processor Reference + Merchant Name
    if has('processor reference') and has('merchant name'):
        # Check settlement gateway signatures too — some files might overlap
        if not has('merchant_ref_number') and not has('order reference'):
            return ('PELPAY', None, None)

    # Cybersource
    if has('merchant_ref_number') and has('amount') and has('merchant_id'):
        # Determine currency
        cur = None
        for col_h in ['currency']:
            if any(col_h == hh for hh in h):
                ci = h.index(col_h)
                # We'll read currency from data later
                cur = col_h
        return ('SETTLEMENT', 'CYBERSOURCE', cur)

    # ChoicePay / MPGS
    if has('order reference') and has('order amount') and has('merchant id'):
        cur = None
        for col_h in ['order amount (currency only)', 'currency']:
            if any(col_h == hh for hh in h):
                cur = col_h
        return ('SETTLEMENT', 'CHOICEPAY', cur)

    # MPGS
    if has('processor reference') and has('settlement amount') and has('merchant code'):
        cur = None
        for col_h in ['currency']:
            if any(col_h == hh for hh in h):
                cur = col_h
        return ('SETTLEMENT', 'MPGS', cur)

    return None

def read_currency_from_file(fpath, cur_col_header):
    """Read currency from first data row of given column."""
    ext = os.path.splitext(fpath)[1].lower()
    try:
        if ext == '.csv':
            with open(fpath, encoding='utf-8-sig', newline='') as fh:
                reader = csv.reader(fh)
                h = [str(c).strip().lower() for c in next(reader, [])]
                if cur_col_header in h:
                    ci = h.index(cur_col_header)
                    for row in reader:
                        if row and len(row) > ci and row[ci]:
                            return str(row[ci]).strip().upper()
                        break
        else:
            wb = openpyxl.load_workbook(fpath, data_only=True)
            ws = wb[wb.sheetnames[0]]
            h = [str(c.value).strip().lower() for c in ws[1]]
            if cur_col_header in h:
                ci = h.index(cur_col_header)
                for row in ws.iter_rows(min_row=2, max_row=2, values_only=True):
                    if row[ci]:
                        return str(row[ci]).strip().upper()
            wb.close()
    except Exception:
        pass
    return None

def currency_from_filename(fpath):
    name = os.path.basename(fpath).upper()
    if 'NGN' in name and 'USD' not in name:
        return 'NGN'
    if 'USD' in name:
        return 'USD'
    return None

# ── Sidebar — date only ──────────────────────────────────
st.sidebar.header('Configuration')
settlement_date = st.sidebar.date_input(
    'Settlement Date',
    value=date.today(),
)

# ── Main — upload everything ─────────────────────────────
st.subheader('Upload Files')
st.markdown('Upload all daily files — the app auto-detects Pelpay vs settlement files and their currencies.')

uploaded_files = st.file_uploader(
    'Choose .xlsx or .csv files',
    type=['xlsx', 'csv'], accept_multiple_files=True,
    label_visibility='collapsed',
)

# ── Process uploads ─────────────────────────────────────-
if uploaded_files:
    pelpay_file = None
    settlements = []  # list of dicts

    for f in uploaded_files:
        ext = os.path.splitext(f.name)[1].lower()
        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=ext if ext else '.xlsx')
        tmp.write(f.getbuffer())
        tpath = tmp.name
        tmp.close()

        try:
            if ext == '.csv':
                with open(tpath, encoding='utf-8-sig', newline='') as fh:
                    reader = csv.reader(fh)
                    hdrs = next(reader, [])
            else:
                wb = openpyxl.load_workbook(tpath, data_only=True)
                ws = wb[wb.sheetnames[0]]
                hdrs = [c.value for c in ws[1]]
                wb.close()

            cls = classify_file(hdrs)
            if cls is None:
                settlements.append({
                    'file': f, 'path': tpath, 'name': f.name,
                    'gateway': None, 'currency': None, 'status': 'UNKNOWN',
                })
                continue

            ftype, gw, cur_col = cls

            if ftype == 'PELPAY':
                if pelpay_file is None:
                    pelpay_file = {'file': f, 'path': tpath, 'name': f.name}
                else:
                    # Duplicate — mark as unknown
                    settlements.append({
                        'file': f, 'path': tpath, 'name': f.name,
                        'gateway': None, 'currency': None,
                        'status': 'DUPLICATE PELPAY (already have one)',
                    })
                continue

            # Settlement file
            if cur_col:
                currency = read_currency_from_file(tpath, cur_col)
            if not currency:
                currency = currency_from_filename(tpath)
            if not currency:
                currency = '?'

            settlements.append({
                'file': f, 'path': tpath, 'name': f.name,
                'gateway': gw, 'currency': currency, 'status': 'OK',
            })
        except Exception as e:
            settlements.append({
                'file': f, 'path': tpath, 'name': f.name,
                'gateway': None, 'currency': None,
                'status': f'Error: {e}',
            })

    # ── File checklist ────────────────────────────────────
    expected = {
        ('CYBERSOURCE', 'NGN'): 'Cybersource NGN',
        ('CYBERSOURCE', 'USD'): 'Cybersource USD',
        ('CHOICEPAY', 'NGN'): 'ChoicePay/MPGS NGN',
        ('CHOICEPAY', 'USD'): 'ChoicePay/MPGS USD',
    }
    detected_settlements = {(s['gateway'], s['currency']) for s in settlements if s['gateway'] and s['status'] == 'OK' and s['currency'] and s['currency'] != '?'}

    st.subheader('Files Required')
    req_cols = st.columns(2)
    with req_cols[0]:
        st.markdown(f"{'✅' if pelpay_file else '❌'} **Pelpay file**")
    with req_cols[1]:
        st.markdown(f"{pelpay_file['name'] if pelpay_file else '— missing —'}")

    for col_idx, (gw_cur, label) in enumerate(expected.items()):
        found = gw_cur in detected_settlements
        with req_cols[col_idx % 2]:
            st.markdown(f"{'✅' if found else '❌'} **{label}**")
            if found:
                names = [s['name'] for s in settlements if s['gateway'] == gw_cur[0] and s['currency'] == gw_cur[1] and s['status'] == 'OK']
                st.caption(', '.join(names))
            else:
                st.caption('— missing —')

    # ── Detection table ───────────────────────────────────
    st.subheader('Detection Results')
    rows = []
    if pelpay_file:
        rows.append({'File': pelpay_file['name'], 'Type': 'Pelpay', 'Gateway': '—', 'Currency': '—', 'Status': 'OK'})
    for s in settlements:
        gw = s['gateway'] if s['gateway'] else '—'
        cu = s['currency'] if s['currency'] else '—'
        rows.append({'File': s['name'], 'Type': 'Settlement' if s['gateway'] else '—', 'Gateway': gw, 'Currency': cu, 'Status': s['status']})
    st.table(rows)

    # ── Validation ───────────────────────────────────────
    errors = []
    if not pelpay_file:
        errors.append('No Pelpay file detected. Must contain: Processor Reference, Merchant Name')
    valid_settlements = [s for s in settlements if s['gateway'] and s['status'] == 'OK' and s['currency'] and s['currency'] != '?']
    unknown = [s for s in settlements if not s['gateway'] and s['status'] not in ('DUPLICATE PELPAY (already have one)',)]
    if unknown:
        errors.append(f'{len(unknown)} file(s) could not be matched. '
                       'Cybersource files need: merchant_ref_number, amount, merchant_id. '
                       'ChoicePay files need: Order Reference, Order Amount, Merchant ID.')
    if len(valid_settlements) < 4:
        errors.append('Need all 4 settlement files (Cybersource NGN/USD + ChoicePay NGN/USD).')
    for s in settlements:
        if s['gateway'] and s['currency'] == '?':
            errors.append(f'Could not determine currency for {s["name"]}. Rename file to include NGN or USD.')

    if errors:
        for e in errors:
            st.warning(e)

    st.divider()

    # ── Run ──────────────────────────────────────────────
    if not errors and pelpay_file:
        if st.button('Run Reconciliation', type='primary'):
            tmpdir = tempfile.mkdtemp()
            try:
                pel_path = pelpay_file['path']
                settle_items = [(s['gateway'], s['currency'], s['path']) for s in valid_settlements]

                date_str = settlement_date.strftime('%Y-%m-%d')
                out_path = os.path.join(tmpdir, f'Reconciliation_{date_str}_MD_Format.xlsx')

                with st.spinner('Running reconciliation…'):
                    result = run(pel_path, settle_items, settlement_date, out_path)

                st.success('Reconciliation complete!')
                c1, c2, c3 = st.columns(3)
                c1.metric('Settlement Rows', result['settle_rows'])
                c2.metric('Matched', result['matched'])
                c3.metric('Missing from Settlement', result['exceptions'])
                st.write('Sheets:', ', '.join(result['sheets']))

                with open(out_path, 'rb') as fh:
                    st.download_button(
                        label='⬇ Download Reconciliation Workbook',
                        data=fh,
                        file_name=os.path.basename(out_path),
                        mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    )
            except Exception as e:
                st.error(f'Reconciliation failed: {e}')
                import traceback; st.code(traceback.format_exc())

else:
    st.info('Upload .xlsx or .csv files to get started.')

# ── Guide ────────────────────────────────────────────────
with st.expander('How to use'):
    st.markdown('''
1. Set the **Settlement Date** in the sidebar.
 2. **Upload all .xlsx / .csv files** — Pelpay + settlement files, all at once.  
   The app auto-detects:
    - **Pelpay file** (needs: Processor Reference, Merchant Name)
    - **Cybersource files** (needs: merchant_ref_number, amount, merchant_id)
    - **ChoicePay files** (needs: Order Reference, Order Amount, Merchant ID)
    - **MPGS files** (needs: Processor Reference, Settlement Amount, Merchant Code)
    - **Currency** from file content or filename (NGN / USD)
3. Review the detection table.
4. Click **Run Reconciliation**.
5. Download the output workbook.
    ''')
